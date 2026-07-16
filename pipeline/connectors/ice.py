"""EIA's ICE electric workbook — national wholesale hub breadth, panel-only.

Keyless XLSX published by EIA (sourced from ICE trade data), one sheet per
year named after the year itself, ~biweekly cadence (observed lag <=8 days).
Every trade date for every hub lives in the same current-year file, so one
fetch emits the full history-to-date for each requested hub; store
value-dedupe makes daily refetches free. Reuses the census.py xlsx parse
convention: header row located by name (never position), target column
located by normalized header text, plausible-range + zero-row drift checks.

Never a splice proxy — cadence (biweekly) and product (trade-weighted
average vs DAM LMP) both differ from the CAISO/MISO daily tail. This
connector only feeds the "power bill" display panel and deep history.

SPIKE-FINAL (docs/superpowers/specs/2026-07-15-power-spike-notes.md §3):
- 11 populated columns; the fourth, `Delivery \nend date`, carries a literal
  embedded newline in its header cell — proof the header row must be
  normalized (whitespace-collapsed) before any name lookup, not just the
  columns this connector actually reads.
- Date cells (`Trade date` et al.) are `datetime.datetime` objects, not
  strings.
- `PJM WH Real Time Peak` is the verified exact PJM Western hub label.
- `ice_ercot_north` does not exist anywhere in this workbook (exhaustively
  searched) — there is no ERCOT/Texas hub of any kind in this source; do not
  invent one. Dropped from Wave 4 scope at the design level, not here.
- Negative wtd-avg prices are real (curtailment hours) — the plausible range
  admits them.
"""
import io
from datetime import datetime

import openpyxl
import requests

from pipeline.connectors.fred import today_et
from pipeline.connectors.util import get_bytes
from pipeline.models import Observation

BASE_URL = "https://www.eia.gov/electricity/wholesale/xls/ice_electric-{year}.xlsx"
HUB_HEADER = "Price hub"
DATE_HEADER = "Trade date"
VALUE_HEADER = "Wtd avg price $/MWh"
PLAUSIBLE = (-100.0, 3000.0)  # $/MWh; negative prices are real (curtailment)


def _norm(cell) -> str:
    return " ".join(str(cell).split()) if cell is not None else ""


def fetch(source_ids: list[str], vintage_date: str | None = None,
          http_get=None, year: str | int | None = None) -> list[Observation]:
    """source_id = the exact ICE hub label (e.g. 'PJM WH Real Time Peak').

    Emits every trade date found for each requested hub. `year` defaults to
    the year of vintage_date (or today) — the workbook is current-year only;
    prior-year history is already in the store by the time the year rolls."""
    http_get = http_get or requests.get
    vintage = vintage_date or today_et()
    yr = year or (vintage_date or today_et())[:4]
    sheet = str(yr)
    url = BASE_URL.format(year=yr)

    wb = openpyxl.load_workbook(
        io.BytesIO(get_bytes(url, http_get)), read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        raise ValueError(
            f"ice {sheet}: expected sheet '{sheet}' not found "
            f"(sheets: {wb.sheetnames}) — structure drift?")
    rows = list(wb[sheet].iter_rows(values_only=True))

    header_i = next((i for i, r in enumerate(rows[:8])
                     if r and _norm(r[0]) == HUB_HEADER), None)
    if header_i is None:
        raise ValueError(f"ice {sheet}: no '{HUB_HEADER}' header row — "
                         "structure drift?")
    header = [_norm(c).lower() for c in rows[header_i]]
    if DATE_HEADER.lower() not in header:
        raise ValueError(f"ice {sheet}: no '{DATE_HEADER}' column — "
                         "structure drift?")
    if VALUE_HEADER.lower() not in header:
        raise ValueError(f"ice {sheet}: no '{VALUE_HEADER}' column — "
                         "structure drift?")
    date_col = header.index(DATE_HEADER.lower())
    value_col = header.index(VALUE_HEADER.lower())

    wanted = {_norm(sid): sid for sid in source_ids}
    counts = {sid: 0 for sid in source_ids}
    out: list[Observation] = []
    for r in rows[header_i + 1:]:
        if not r:
            continue
        sid = wanted.get(_norm(r[0]))
        if sid is None:
            continue  # not a requested hub — includes footer/note rows
        date_cell = r[date_col] if date_col < len(r) else None
        if not isinstance(date_cell, datetime):
            continue  # not a real trade-date row (e.g. footer text)
        raw = r[value_col] if value_col < len(r) else None
        if raw is None or str(raw).strip() == "":
            continue  # blank price cell skipped
        value = float(raw)
        if not PLAUSIBLE[0] <= value <= PLAUSIBLE[1]:
            raise ValueError(
                f"ice {sid}: value {value} outside plausible range "
                f"{PLAUSIBLE} — structure drift?")
        counts[sid] += 1
        out.append(Observation(series_code=sid, obs_date=date_cell.date().isoformat(),
                               value=value, vintage_date=vintage,
                               source="ICE", route="XLSX"))

    missing = [sid for sid in source_ids if counts[sid] == 0]
    if missing:
        raise ValueError(f"ice: zero rows for hub(s) {missing} — "
                         "structure drift?")
    return out
