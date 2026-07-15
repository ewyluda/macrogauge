"""Census C30 'Value of Construction Put in Place' workbooks — data center column.

Census retired the EITS timeseries API (verified HTTP 302, 2026-07-15), so the
published xlsx workbooks are the only programmatic route. Each fetch carries
the full 2014->now history; vintage.append's value-dedupe means only genuine
revisions (Census p/r cycles) write new rows — an auditable revision trail.

Drift protection (house convention, xlsx dialect): pinned sheet name per file,
header row located by 'Date', target column located by header TEXT (never
position), strict date-label regex, plausible value range — any miss raises a
"structure drift?" ValueError contained by collect's isolation boundary.
"""
import io
import re
from datetime import datetime

import openpyxl
import requests

from pipeline.connectors.fred import today_et
from pipeline.connectors.util import get_bytes
from pipeline.models import Observation

BASE_URL = "https://www.census.gov/construction/c30/xlsx/"
SHEETS = {"privsatime.xlsx": "Private SA", "privtime.xlsx": "Private NSA"}
DATE_RE = re.compile(r"^[A-Z][a-z]{2}-\d{2}[pr]?$")
VALUE_RANGE = (50.0, 500_000.0)  # $M; NSA runs ~124->5,059, SAAR ~1,500->61,000


def _norm(cell) -> str:
    return " ".join(str(cell).split()) if cell is not None else ""


def _parse_date(label: str) -> str:
    # 'May-26p' / 'Apr-26r' -> '2026-05-01' (p/r are Census revision suffixes)
    return datetime.strptime(label.rstrip("pr"), "%b-%y").strftime("%Y-%m-01")


def _column(rows: list, filename: str, column: str) -> dict[str, float]:
    header_i = next((i for i, r in enumerate(rows[:8])
                     if r and _norm(r[0]) == "Date"), None)
    if header_i is None:
        raise ValueError(f"census {filename}: no 'Date' header row — structure drift?")
    header = [_norm(c).lower() for c in rows[header_i]]
    if column.lower() not in header:
        raise ValueError(f"census {filename}: no '{column}' column — structure drift?")
    col = header.index(column.lower())
    out: dict[str, float] = {}
    for r in rows[header_i + 1:]:
        label = _norm(r[0]) if r else ""
        if not DATE_RE.match(label):
            break  # footer disclosure notes end the data block
        if col < len(r) and r[col] is not None and str(r[col]).strip() != "":
            value = float(r[col])
            if not VALUE_RANGE[0] <= value <= VALUE_RANGE[1]:
                raise ValueError(
                    f"census {filename}: {column} value {value} outside plausible "
                    f"range {VALUE_RANGE} — structure drift?")
            out[_parse_date(label)] = value
    if not out:
        raise ValueError(f"census {filename}: zero data rows parsed — structure drift?")
    return out


def fetch(source_ids: list[str], vintage_date: str | None = None,
          http_get=None) -> list[Observation]:
    """source_id format: '<filename>:<column header>'."""
    http_get = http_get or requests.get
    vintage = vintage_date or today_et()
    wanted: dict[str, list[tuple[str, str]]] = {}
    for sid in source_ids:
        filename, column = sid.split(":", 1)
        wanted.setdefault(filename, []).append((sid, column))
    out: list[Observation] = []
    for filename, cols in wanted.items():
        wb = openpyxl.load_workbook(
            io.BytesIO(get_bytes(BASE_URL + filename, http_get)), read_only=True)
        expected = SHEETS.get(filename)
        if expected is None or expected not in wb.sheetnames:
            raise ValueError(
                f"census {filename}: expected sheet '{expected}' not found "
                f"(sheets: {wb.sheetnames}) — structure drift?")
        rows = list(wb[expected].iter_rows(values_only=True))
        for sid, column in cols:
            for obs_date, value in _column(rows, filename, column).items():
                out.append(Observation(series_code=sid, obs_date=obs_date,
                                       value=value, vintage_date=vintage,
                                       source="CENSUS", route="XLSX"))
    return out
